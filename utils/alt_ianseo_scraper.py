import os
import re
from pathlib import Path

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from PyQt6.QtCore import QThread, pyqtSignal


# ---------------------------------------------------------------------------
# Regex patterns
# ---------------------------------------------------------------------------

ILLEGAL_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')
FORBIDDEN_FILE_CHARS = re.compile(r'[\\/:*?"<>|\x00]')
YEAR_PATTERN = re.compile(r'/(\d{4})/')
UNDER_PATTERN = re.compile(r'(under)\s?(\d{2}\b)', re.IGNORECASE)
U_PATTERN = re.compile(r'u\s?(\d{2}\b)', re.IGNORECASE)

SEX_PATTERNS = re.compile(
    r'(?P<male>(?<!o)men|(?<!e)male)'
    r'|(?P<open>open)'
    r'|(?P<female>women|female|ladies)'
    r'|(?P<enby>non\s?\-?binary|\bnb\b)',
    re.IGNORECASE
)

AGE_PATTERNS = re.compile(
    r'(?P<fiftyplus>over\s?50|50\s?\+)'
    r'|(?P<junior>junior|\bjun\b)'
    r'|(?P<underxx>under\s?\d{2}\b)'
    r'|(?P<uxx>u\s?\d{2}\b)',
    re.IGNORECASE
)

STYLE_PATTERNS = re.compile(r'compound|recurve|barebow|longbow', re.IGNORECASE)
ILLEGAL_STYLE_PATTERNS = re.compile(
    r'recurve\s?/?\s?barebow|barebow\s?/?\s?recurve|flatbow|asiatic|compound barebow|traditional',
    re.IGNORECASE
)

STYLE_DICT = {
    "compound": "C",
    "recurve":  "R",
    "barebow":  "B",
    "longbow":  "L"
}

# ---------------------------------------------------------------------------
# Name standardisation
# ---------------------------------------------------------------------------

def generate_standard_name(sheet_header: str) -> str:
    """
    Convert a longform table name into a standard format
    similar to how IANSEO formats table names (e.g. IQRM).
    Returns "remove" if the header cannot be standardised.
    """
    if ILLEGAL_STYLE_PATTERNS.search(sheet_header):
        return "remove"

    sex = ""
    age = ""
    style = ""

    sex_match = SEX_PATTERNS.search(sheet_header)
    if sex_match:
        group_name = sex_match.lastgroup
        if group_name == "male":
            sex = "M"
        elif group_name == "open":
            sex = "O"
        elif group_name == "female":
            sex = "W"
        elif group_name == "enby":
            sex = "NB"

    age_match = AGE_PATTERNS.search(sheet_header)
    if age_match:
        group_name = age_match.lastgroup
        group_value = age_match.group(group_name)
        if group_name == "fiftyplus":
            age = "50+"
        elif group_name == "junior":
            age = "J"
        elif group_name == "underxx":
            age = UNDER_PATTERN.sub(r'U\2', group_value)
        elif group_name == "uxx":
            age = U_PATTERN.sub(r'U\1', group_value)

    style_match = STYLE_PATTERNS.search(sheet_header)
    if style_match:
        style = STYLE_DICT[style_match.group().lower()]

    if not sex or not style:
        return "remove"

    return f"IQ{sex}{age}{style}"

# ---------------------------------------------------------------------------
# HTML parsing helpers
# ---------------------------------------------------------------------------

def get_cell_text(cell) -> str:
    return cell.get_text(separator=" ", strip=True)


def is_section_title_row(tr) -> bool:
    """
    Returns True if a <tr> is a section title row:
    a single <th> with a large colspan and no subheader class.
    """
    ths = tr.find_all("th", recursive=False)
    if len(ths) == 1:
        colspan = ths[0].get("colspan")
        if colspan:
            try:
                return int(colspan) > 3 and "slim-subheaders" not in tr.get("class", [])
            except ValueError:
                pass
    return False


def is_column_header_row(tr) -> bool:
    """
    Returns True if a <tr> contains the actual column headers.
    Identified by slim-subheaders class (page2) or multiple <th> elements (page1).
    """
    ths = tr.find_all("th", recursive=False)
    if not ths:
        return False
    if "slim-subheaders" in tr.get("class", []):
        return True
    return len(ths) > 1


def is_secondary_line(tr) -> bool:
    """Returns True if a <tr> is a secondary detail line to be skipped."""
    return tr.get("id") == "secondary-lines"

# ---------------------------------------------------------------------------
# Section extraction
# ---------------------------------------------------------------------------

def extract_sections_from_interleaved_table(table) -> list[dict]:
    """
    Handles page2-style tables where one <table> contains multiple
    interleaved <thead>/<tbody> pairs. Each <thead> with a section title
    row starts a new logical section.
    """
    sections = []
    current_section = None

    for block in table.find_all(["thead", "tbody"], recursive=False):
        if block.name == "thead":
            section_title = None
            col_headers = []

            for tr in block.find_all("tr", recursive=False):
                if is_section_title_row(tr):
                    section_title = get_cell_text(tr.find_all("th", recursive=False)[0])
                elif is_column_header_row(tr):
                    col_headers = [get_cell_text(th) for th in tr.find_all("th", recursive=False)]

            if section_title:
                if current_section is not None:
                    sections.append(current_section)
                current_section = {
                    "section_header": section_title,
                    "col_headers": col_headers,
                    "rows": []
                }
            elif col_headers and current_section is not None:
                current_section["col_headers"] = col_headers

        elif block.name == "tbody":
            if current_section is None:
                current_section = {"section_header": None, "col_headers": [], "rows": []}

            for tr in block.find_all("tr", recursive=False):
                if is_secondary_line(tr):
                    continue
                tds = tr.find_all("td", recursive=False)
                if not tds:
                    continue
                row_data = [get_cell_text(td) for td in tds]
                if any(cell and cell != "\xa0" for cell in row_data):
                    current_section["rows"].append(row_data)

    if current_section is not None and current_section["rows"]:
        sections.append(current_section)

    return sections


def extract_sections_from_accordion_table(table, accordion_div) -> dict:
    """
    Handles page1-style tables where each table lives inside an accordion div.
    The section header comes from the accordion's title div.
    """
    title_div = accordion_div.find("div", class_="title", recursive=False)
    section_header = None
    if title_div:
        span = title_div.find("span")
        section_header = span.get_text(strip=True) if span else title_div.get_text(strip=True)

    col_headers = []
    rows = []

    for tr in table.find_all("tr"):
        ths = tr.find_all("th", recursive=False)
        tds = tr.find_all("td", recursive=False)
        if ths and not col_headers:
            col_headers = [get_cell_text(th) for th in ths]
        elif tds:
            row_data = [get_cell_text(td) for td in tds]
            if any(cell and cell != "\xa0" for cell in row_data):
                rows.append(row_data)

    return {"section_header": section_header, "col_headers": col_headers, "rows": rows}

# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def write_section_to_sheet(ws, section: dict, comp_name: str, year: str) -> None:
    """Write a single section dict to an already-created worksheet."""
    current_row = 1

    if section["col_headers"]:
        ws.cell(row=current_row, column=1, value="Year")
        ws.cell(row=current_row, column=2, value="Competition")
        for col_idx, header in enumerate(section["col_headers"], start=3):
            ws.cell(row=current_row, column=col_idx, value=header)
        current_row += 1

    for row_data in section["rows"]:
        ws.cell(row=current_row, column=1, value=year)
        ws.cell(row=current_row, column=2, value=comp_name)
        for col_idx, cell_value in enumerate(row_data, start=3):
            ws.cell(row=current_row, column=col_idx, value=cell_value)
        current_row += 1


def tables_to_excel(url: str, output_file: str, comp_name: str, year: str) -> None:
    """
    Fetches the page at `url`, detects its format, extracts all result
    sections, and writes each as a separate Excel sheet.
    """
    soup = BeautifulSoup(fetch_page(url), "html.parser")
    all_sections = []

    accordion_divs = soup.find_all("div", class_="accordion")
    if accordion_divs:
        for acc_div in accordion_divs:
            table = acc_div.find("table")
            if table:
                section = extract_sections_from_accordion_table(table, acc_div)
                if section["rows"]:
                    all_sections.append(section)

    if not all_sections:
        for table in soup.find_all("table"):
            if len(table.find_all("thead", recursive=False)) > 1:
                sections = extract_sections_from_interleaved_table(table)
                all_sections.extend(s for s in sections if s["rows"])

    if not all_sections:
        #print("No result sections found on the page.")
        return

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    sheet_name_counts = {}

    for section in all_sections:
        raw_name = section["section_header"] or "Table"
        sheet_name = generate_standard_name(raw_name)

        if sheet_name == "remove":
            continue

        if sheet_name in sheet_name_counts:
            sheet_name_counts[sheet_name] += 1
            sheet_name = f"{sheet_name}_{sheet_name_counts[sheet_name]}"
        else:
            sheet_name_counts[sheet_name] = 1

        ws = workbook.create_sheet(title=sheet_name)
        write_section_to_sheet(ws, section, comp_name, year)

    if not workbook.sheetnames:
        #print("No sheets to save after filtering.")
        return

    workbook.save(output_file)
    #print(f"Saved {len(workbook.sheetnames)} sheet(s) to '{output_file}'.")

# ---------------------------------------------------------------------------
# Page scraping helpers
# ---------------------------------------------------------------------------

def fetch_page(url: str) -> str:
    """Fetch a webpage into RAM and return its HTML content as a string."""
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.text


def get_ic_link(url: str) -> str | None:
    """
    Fetches a webpage into RAM and returns the href of the link ending with IC.php.
    Returns None if no such link is found.
    """
    soup = BeautifulSoup(fetch_page(url), "html.parser")
    for tag in soup.find_all("a", href=True):
        if tag["href"].endswith("IC.php"):
            return f"https://www.ianseo.net{tag["href"]}"
    return None


def get_comp_name(url: str) -> str:
    """
    Get competition name from a given Ianseo page and sanitise it.
    Returns a sanitised competition name with spaces replaced by underscores.
    """
    soup = BeautifulSoup(fetch_page(url), "html.parser")

    center = soup.select_one(".results-header-center")
    if center is not None:
        raw_name = center.find("div").get_text(strip=True)
    else:
        header = soup.find("table", id="TourHeader")
        th = header.find("th") if header else None
        raw_name = th.get_text(separator="\n", strip=True).split("\n")[0]

    return re.sub(FORBIDDEN_FILE_CHARS, "_", raw_name.replace(" ", "_"))


class AltIanseoScraper(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal()

    def __init__(self, urls, parent=None):
        super(AltIanseoScraper, self).__init__(parent)
        self.urls = urls

    def run(self):
        save_directory = Path("temp/alt_ianseo")
        save_directory.mkdir(parents=True, exist_ok=True)
        i = 0
        self.progress.emit(i)
        self.totalUrls = len(self.urls)

        for url in self.urls:
            comp_link = get_ic_link(url)
            if not comp_link:
                continue

            comp_name = get_comp_name(url)
            year = re.search(YEAR_PATTERN, comp_link).group(1)
            file_name = f"{comp_name[:26]}_{year}.xlsx"

            if file_name in os.listdir(save_directory):
                file_name = "1" + file_name[1:]

            tables_to_excel(comp_link, save_directory / file_name, comp_name, year)

            i += 1
            self.progress.emit(i)

        self.finished.emit()